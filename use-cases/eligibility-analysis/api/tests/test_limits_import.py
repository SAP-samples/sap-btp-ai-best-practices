from __future__ import annotations

import io
import unittest
from unittest.mock import patch

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.optimizer.model.limits_import import (
    LimitsImportError,
    import_limits_payload,
)
from app.routers.optimizer import router as optimizer_router
from app.security import get_api_key


def _build_limits_workbook_bytes(
    rows: list[list[object]],
    *,
    headers: list[str] | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    columns = headers or [
        "ID Seller",
        "Group Debtor",
        "ID Debtor",
        "Seller Limit",
        "Currency Seller LM",
        "Group Limit",
        "Currency Group LM",
        "Debtor Limit",
        "Currency Debtor LM ",
        "Facility Limit",
        "Currency Facility",
    ]
    for col_idx, value in enumerate(columns, start=1):
        sheet.cell(row=1, column=col_idx, value=value)
    for row_idx, row_values in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class TestLimitsImportParser(unittest.TestCase):
    def test_excel_import_mixed_currency_and_group_normalization(self) -> None:
        content = _build_limits_workbook_bytes(
            [
                ["CC01", 1322.0, 1000001717, None, None, 10_000_000, "EUR", 7_000_000, "EUR", 750_000_000, "EUR"],
                ["CC02", 4399.0, 1000001455, None, None, 37_000_000, "GBP", 12_850_000, "GBP", 750_000_000, "EUR"],
                ["CC03", None, 1000002001, 50_000_000, "GBP", None, None, 1_000_000, "GBP", None, None],
            ]
        )

        payload, summary = import_limits_payload(content, filename="limits.xlsx", existing_limits={})
        self.assertEqual(payload["synthetic_generation"]["enabled"], False)
        self.assertEqual(payload["customer_to_group"]["1000001717"], "1322")
        self.assertEqual(payload["customer_to_group"]["1000001455"], "4399")
        self.assertNotIn("1000002001", payload["customer_to_group"])

        self.assertEqual(payload["facility_limits_by_company_code"]["CC01"], 750000000.0)
        self.assertEqual(payload["facility_limits_by_company_code"]["CC02"], 750000000.0)
        self.assertEqual(payload["facility_limits_by_company_code"]["CC03"], round(50_000_000 / 0.87, 2))

        self.assertEqual(payload["group_limits"]["1322"], 10000000.0)
        self.assertEqual(payload["group_limits"]["4399"], round(37_000_000 / 0.87, 2))
        self.assertEqual(payload["customer_limits"]["1000001455"], round(12_850_000 / 0.87, 2))
        self.assertEqual(payload["customer_limits"]["1000002001"], round(1_000_000 / 0.87, 2))
        self.assertGreaterEqual(summary["gbp_conversions"], 4)

    def test_excel_import_rejects_unsupported_currency(self) -> None:
        content = _build_limits_workbook_bytes(
            [["CC01", None, 1001, None, None, None, None, 500, "USD", 1000, "EUR"]]
        )
        with self.assertRaises(LimitsImportError):
            import_limits_payload(content, filename="limits.xlsx", existing_limits={})

    def test_excel_import_detects_conflicting_group_limit(self) -> None:
        content = _build_limits_workbook_bytes(
            [
                ["CC01", 1322.0, 1001, None, None, 10_000, "EUR", 300, "EUR", 100_000, "EUR"],
                ["CC01", 1322.0, 1002, None, None, 12_000, "EUR", 200, "EUR", 100_000, "EUR"],
            ]
        )
        with self.assertRaises(LimitsImportError):
            import_limits_payload(content, filename="limits.xlsx", existing_limits={})


class _ManagerStub:
    def get_limits(self, process_id: str):
        if process_id != "proc-1":
            raise ValueError("not found")
        return {
            "defaults": {
                "customer_limit_fraction_of_facility": 0.2,
                "group_limit_fraction_of_facility": 0.3,
            }
        }


class TestLimitsImportApi(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.app = FastAPI()
        cls.app.include_router(optimizer_router, prefix="/api", tags=["optimizer"])
        cls.app.dependency_overrides[get_api_key] = lambda: "test"
        cls.client = TestClient(cls.app)

    @classmethod
    def tearDownClass(cls) -> None:
        cls.app.dependency_overrides.pop(get_api_key, None)

    @patch("app.routers.optimizer.get_process_manager")
    def test_import_limits_endpoint_excel(self, mock_get_mgr) -> None:
        mock_get_mgr.return_value = _ManagerStub()
        content = _build_limits_workbook_bytes(
            [["CC01", 1322.0, 1001, None, None, 10_000, "EUR", 300, "EUR", 100_000, "EUR"]]
        )

        response = self.client.post(
            "/api/optimizer/processes/proc-1/limits/import",
            files={
                "file": (
                    "limits.xlsx",
                    content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        self.assertEqual(response.status_code, 200, response.text)
        body = response.json()
        self.assertIn("limits_payload", body)
        self.assertIn("import_summary", body)
        self.assertEqual(body["limits_payload"]["synthetic_generation"]["enabled"], False)
        self.assertEqual(
            body["limits_payload"]["defaults"]["customer_limit_fraction_of_facility"],
            0.2,
        )

    @patch("app.routers.optimizer.get_process_manager")
    def test_import_limits_endpoint_invalid_excel_schema(self, mock_get_mgr) -> None:
        mock_get_mgr.return_value = _ManagerStub()
        content = _build_limits_workbook_bytes([["x", "y"]], headers=["A", "B"])
        response = self.client.post(
            "/api/optimizer/processes/proc-1/limits/import",
            files={
                "file": (
                    "limits.xlsx",
                    content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        self.assertEqual(response.status_code, 400, response.text)
        self.assertIn("Missing required columns", response.json().get("detail", ""))


if __name__ == "__main__":
    unittest.main()
