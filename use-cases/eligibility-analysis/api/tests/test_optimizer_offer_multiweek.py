import importlib.util
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from app.optimizer.io.load_extraction import load_extraction
from app.optimizer.io.load_offer_file import load_offer_file
from app.optimizer.model.lifecycle import derive_lifecycle
from app.optimizer.model.limits import ResolvedLimits
from app.optimizer.opt.explain_multi_week import explain_multi_week_non_selection
from app.optimizer.opt.optimizer_multi_week import (
    MultiWeekOptimizerSettings,
    optimize_multi_week,
)

HAS_ORTOOLS = importlib.util.find_spec("ortools") is not None


def _write_summary_funded_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Funded Invoices"

    ws["A1"] = "OFFER SUMMARY"
    ws["A4"] = "Offer ID:"
    ws["B4"] = 2
    ws["A5"] = "Client Name:"
    ws["B5"] = "COMPANY CODE"
    ws["A6"] = "Program Name:"
    ws["B6"] = "DEVELOPMENT SYSTEM"
    ws["A7"] = "Funding date:"
    ws["B7"] = pd.Timestamp("2025-01-28")
    ws["A8"] = "Purchased Invoices"

    headers = [
        "Seller Name",
        "Seller Client ID",
        "Debtor Name",
        "Debtor Client ID",
        "Doc Type",
        "Doc Ref",
        "Issue Date",
        "Due Date",
        "Reconciliation Date",
        "Purchase Date",
        "Original CCY",
        "Total Invoice Amount (O. CCY)",
        "Total Net Amount (O. CCY)",
        "Eligible Amount (O. CCY)",
        "Funding CCY",
        "Total Invoice Amount (F. CCY)",
        "Eligible amount (F. CCY)",
        "Base Rate",
        "Credit/flat fee",
        "Margin",
        "All in rate",
        "Interest",
        "Purchase Amount",
    ]
    for col_idx, value in enumerate(headers, start=1):
        ws.cell(row=11, column=col_idx, value=value)

    rows = [
        [
            "COMPANY 1", "2410", "Customer A", "CUST-A", "I", "DOC-1",
            pd.Timestamp("2026-01-27"), pd.Timestamp("2026-02-26"), pd.Timestamp("2026-03-05"),
            pd.Timestamp("2025-01-28"), "EUR", 2500, 2500, 2500, "EUR", 2500, 2500,
            3.490917, None, 1.0, 4.490917, 8.3, 2500,
        ],
        [
            "COMPANY 1", "2410", "Customer B", "CUST-B", "I", "DOC-2",
            pd.Timestamp("2026-01-26"), pd.Timestamp("2026-11-24"), pd.Timestamp("2026-12-01"),
            pd.Timestamp("2025-01-28"), "USD", 100, 100, 100, "USD", 100, 100,
            3.490917, None, 1.0, 4.490917, 2.76, 100,
        ],
        [
            None, None, None, None, None, None,
            None, None, None, None, "Totals:", 2600, 2600, 2600, None, 2600, 2600,
            None, None, None, None, 11.06, 2600,
        ],
    ]
    for row_idx, row_values in enumerate(rows, start=12):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(path)


class TestOfferAndExtractionMapping(unittest.TestCase):
    def test_offer_loader_maps_canonical_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "offer.xlsx"
            df = pd.DataFrame(
                {
                    "PROGRAMA": ["DEV"],
                    "ID SELLER": ["2410"],
                    "ID DEBTOR": ["CUST01"],
                    "INVOICE REF": ["INV-1"],
                    "TOTAL NET VALUE (ORIGINAL CCY)": [1200.5],
                    "TOTAL INVOICE AMOUNT (ORIGINAL CCY)": [1300.0],
                    "ORIGINAL CURRENCY": ["EUR"],
                    "FUNDING CURRENCY": ["EUR"],
                    "EXCHANGE RATE": [1.0],
                    "ISSUANCE DATE": ["2026-01-10"],
                    "DESPATCH DATE": ["2026-01-11"],
                    "DUE DATE": ["2026-02-10"],
                    "MARGIN": [2.1],
                }
            )
            df.to_excel(path, sheet_name="Sheet1", index=False)

            loaded, report = load_offer_file(path, sheet_name="Sheet1")
            self.assertEqual(report.loaded_rows, 1)
            self.assertIn("invoice_reference", loaded.columns)
            self.assertIn("candidate_amount", loaded.columns)
            self.assertIn("Invoice Reference", loaded.columns)
            self.assertAlmostEqual(float(loaded.iloc[0]["candidate_amount"]), 1200.5, places=2)

    def test_extraction_loader_maps_credit_release_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "extraction.xlsx"
            df = pd.DataFrame(
                {
                    "Offer File Date (UTC)": ["2026-01-07T10:00:00"],
                    "Summary File Date (UTC)": ["2026-01-08T11:00:00"],
                    "Reconciliation File Date (UTC)": ["2026-01-30T13:00:00"],
                    "Paid On (Europe, Madrid)": ["2026-02-01"],
                    "Repurchase Date": [None],
                    "Repurchase": [None],
                    "Company Code": ["Company Code 1"],
                    "Customer": ["100456043"],
                    "Invoice Reference": ["INV-1"],
                    "Document Number": ["DOC-1"],
                    "Due Date": ["2026-02-15"],
                    "Purchase Price": [1000.0],
                    "Status": ["Paid to Lender"],
                    "Reason": [None],
                }
            )
            df.to_excel(path, sheet_name="SAPUI5 Export", index=False)

            loaded, report = load_extraction(path, sheet_name="SAPUI5 Export")
            self.assertEqual(report.loaded_rows, 1)
            self.assertIn("credit_start", loaded.columns)
            self.assertIn("credit_release", loaded.columns)
            self.assertTrue(pd.notna(loaded.iloc[0]["credit_release"]))

    def test_extraction_loader_supports_summary_funded_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "summary.xlsx"
            _write_summary_funded_workbook(path)

            loaded, report = load_extraction(path, sheet_name="Funded Invoices")
            self.assertEqual(report.sheet_name, "Funded Invoices")
            self.assertEqual(report.loaded_rows, 2)
            self.assertTrue((loaded["Status"] == "Accepted").all())
            self.assertEqual(loaded.iloc[0]["Company Code"], "2410")
            self.assertEqual(loaded.iloc[0]["Invoice Reference"], "DOC-1")
            self.assertEqual(loaded.iloc[0]["Document Number"], "DOC-1")
            self.assertEqual(float(loaded.iloc[0]["Purchase Price"]), 2500.0)
            self.assertEqual(
                str(pd.to_datetime(loaded.iloc[0]["Offer File Date (UTC)"]).date()),
                "2025-01-28",
            )
            self.assertNotIn("Totals:", set(loaded["Currency"].dropna().astype(str)))

    def test_extraction_loader_falls_back_from_default_to_summary_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "summary.xlsx"
            _write_summary_funded_workbook(path)

            loaded, report = load_extraction(path, sheet_name="SAPUI5 Export")
            self.assertEqual(report.sheet_name, "Funded Invoices")
            self.assertEqual(report.loaded_rows, 2)
            self.assertSetEqual(
                set(loaded["Invoice Reference"].tolist()),
                {"DOC-1", "DOC-2"},
            )


class TestLifecycleReleaseSemantics(unittest.TestCase):
    def test_default_release_is_reconciliation_file_date(self) -> None:
        df = pd.DataFrame(
            {
                "Summary File Date (UTC)": ["2026-01-08"],
                "Reconciliation File Date (UTC)": ["2026-02-01"],
                "Reconciliation Date": ["2026-01-28"],
                "Repurchase Date": ["2026-01-20"],
                "Paid On (Europe, Madrid)": ["2026-02-10"],
            }
        )
        out = derive_lifecycle(df)
        self.assertEqual(pd.to_datetime(out.iloc[0]["credit_end"]), pd.Timestamp("2026-02-01"))
        self.assertNotEqual(pd.to_datetime(out.iloc[0]["credit_end"]), pd.Timestamp("2026-01-20"))


@unittest.skipUnless(HAS_ORTOOLS, "ortools is required for multi-week optimizer tests")
class TestMultiWeekOptimizer(unittest.TestCase):
    def _sample_candidates(self) -> pd.DataFrame:
        return pd.DataFrame(
            {
                "Invoice Reference": ["INV-1", "INV-2", "INV-3"],
                "Company Code": ["C1", "C1", "C1"],
                "Customer": ["U1", "U1", "U1"],
                "Purchase Price": [40.0, 40.0, 40.0],
                "Due Date": [pd.Timestamp("2026-03-31")] * 3,
                "expected_lifetime_weeks": [2, 2, 2],
            }
        )

    def _limits(self) -> ResolvedLimits:
        return ResolvedLimits(
            facility_limits_by_company_code={"C1": 10000},
            customer_limits={"U1": 10000},
            group_limits={},
            customer_to_group={},
        )

    def test_weekly_constraints_respected(self) -> None:
        candidates = self._sample_candidates()
        weeks = [pd.Timestamp("2026-02-02"), pd.Timestamp("2026-02-09"), pd.Timestamp("2026-02-16")]
        base = {
            pd.Timestamp("2026-02-02"): {"facility": {"C1": 2000}, "customer": {"U1": 2000}},
            pd.Timestamp("2026-02-09"): {"facility": {"C1": 2000}, "customer": {"U1": 2000}},
            pd.Timestamp("2026-02-16"): {"facility": {"C1": 2000}, "customer": {"U1": 2000}},
        }
        settings = MultiWeekOptimizerSettings(horizon_weeks=3, attempt_cap=1, default_lifetime_weeks=2)
        result = optimize_multi_week(candidates, self._limits(), weeks, base, settings)

        for week, entities in result.facility_weekly_usage.items():
            for _, usage in entities.items():
                self.assertLessEqual(usage["used_total"], usage["limit"] + 1e-6)
        for week, entities in result.customer_weekly_usage.items():
            for _, usage in entities.items():
                self.assertLessEqual(usage["used_total"], usage["limit"] + 1e-6)

    def test_non_selected_invoices_have_reason_codes(self) -> None:
        candidates = self._sample_candidates()
        weeks = [pd.Timestamp("2026-02-02"), pd.Timestamp("2026-02-09")]
        limits = ResolvedLimits(
            facility_limits_by_company_code={"C1": 4000},
            customer_limits={"U1": 4000},
            group_limits={},
            customer_to_group={},
        )
        result = optimize_multi_week(
            candidates,
            limits,
            weeks,
            base_weekly_exposure={},
            settings=MultiWeekOptimizerSettings(horizon_weeks=2, attempt_cap=1, default_lifetime_weeks=2),
        )
        explained = explain_multi_week_non_selection(candidates, result, limits).explained_not_selected_df
        if not explained.empty:
            self.assertIn("excluded_reason", explained.columns)
            allowed = {
                "EXPIRED_WINDOW",
                "FACILITY_CAP_BINDING",
                "CUSTOMER_CAP_BINDING",
                "GROUP_CAP_BINDING",
                "DEFERRED_FOR_CAPACITY",
            }
            self.assertTrue(set(explained["excluded_reason"]).issubset(allowed))

    def test_offer_week_blocks_early_submission(self) -> None:
        candidates = pd.DataFrame(
            {
                "Invoice Reference": ["INV-A", "INV-B"],
                "Company Code": ["C1", "C1"],
                "Customer": ["U1", "U1"],
                "Purchase Price": [40.0, 40.0],
                "Due Date": [pd.Timestamp("2026-03-31"), pd.Timestamp("2026-03-31")],
                "Offer File Date (UTC)": [pd.Timestamp("2026-02-03"), pd.Timestamp("2026-02-10")],
                "expected_lifetime_weeks": [1, 1],
            }
        )
        weeks = [pd.Timestamp("2026-02-03"), pd.Timestamp("2026-02-10"), pd.Timestamp("2026-02-17")]
        settings = MultiWeekOptimizerSettings(horizon_weeks=3, attempt_cap=1, default_lifetime_weeks=1)
        result = optimize_multi_week(candidates, self._limits(), weeks, base_weekly_exposure={}, settings=settings)
        self.assertFalse(result.weekly_plan_df.empty)

        planned_b = result.weekly_plan_df[
            result.weekly_plan_df["Invoice Reference"] == "INV-B"
        ]["planned_week_start"]
        self.assertTrue(planned_b.notna().all())
        self.assertTrue((planned_b >= pd.Timestamp("2026-02-10")).all())


if __name__ == "__main__":
    unittest.main()
