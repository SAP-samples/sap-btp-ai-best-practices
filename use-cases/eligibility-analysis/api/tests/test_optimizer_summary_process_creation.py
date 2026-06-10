import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.services.optimizer import process_manager as process_manager_module
from app.services.optimizer.process_manager import ProcessManager


class _CreateStoreStub:
    def __init__(self) -> None:
        self.records = {}

    def create_process(
        self,
        process_id: str,
        process_dir: str,
        extraction_filename: str,
        cohort: str | None = None,
        sheet_name: str = "SAPUI5 Export",
        source_profile: str = "extraction_file",
    ):
        record = {
            "id": process_id,
            "status": "created",
            "process_dir": process_dir,
            "extraction_filename": extraction_filename,
            "cohort": cohort,
            "sheet_name": sheet_name,
            "source_profile": source_profile,
        }
        self.records[process_id] = record
        return dict(record)


def _write_summary_funded_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Funded Invoices"

    ws["A1"] = "OFFER SUMMARY"
    ws["A4"] = "Offer ID:"
    ws["B4"] = 2
    ws["A7"] = "Funding date:"
    ws["B7"] = "2025-01-28"
    ws["A8"] = "Purchased Invoices"

    headers = [
        "Seller Name",
        "Seller Client ID",
        "Debtor Name",
        "Debtor Client ID",
        "Doc Type",
        "Doc Ref",
        "Issue Date",
        "Due Date",
        "Reconciliation Date",
        "Purchase Date",
        "Original CCY",
        "Total Invoice Amount (O. CCY)",
        "Total Net Amount (O. CCY)",
        "Eligible Amount (O. CCY)",
        "Funding CCY",
        "Total Invoice Amount (F. CCY)",
        "Eligible amount (F. CCY)",
        "Base Rate",
        "Credit/flat fee",
        "Margin",
        "All in rate",
        "Interest",
        "Purchase Amount",
    ]
    for col_idx, value in enumerate(headers, start=1):
        ws.cell(row=11, column=col_idx, value=value)

    rows = [
        [
            "COMPANY 1", "2410", "Customer A", "CUST-A", "I", "DOC-1",
            "2026-01-27", "2026-02-26", "2026-03-05", "2025-01-28",
            "EUR", 2500, 2500, 2500, "EUR", 2500, 2500, 3.490917, None, 1.0, 4.490917, 8.3, 2500,
        ],
        [
            "COMPANY 1", "2410", "Customer B", "CUST-B", "I", "DOC-2",
            "2026-01-26", "2026-11-24", "2026-12-01", "2025-01-28",
            "USD", 100, 100, 100, "USD", 100, 100, 3.490917, None, 1.0, 4.490917, 2.76, 100,
        ],
    ]
    for row_idx, row_values in enumerate(rows, start=12):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(path)


class TestSummaryProcessCreation(unittest.TestCase):
    def test_create_process_resolves_summary_sheet_and_detects_cohorts(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "summary.xlsx"
            _write_summary_funded_workbook(workbook_path)
            content = workbook_path.read_bytes()

            store = _CreateStoreStub()
            manager = ProcessManager(store=store)

            original_data_dir = process_manager_module.DATA_DIR
            process_manager_module.DATA_DIR = Path(tmpdir) / "optimizer_runs"
            try:
                record = manager.create_process(
                    file_content=content,
                    filename="summary.xlsx",
                    cohort=None,
                    sheet_name="SAPUI5 Export",
                )
            finally:
                process_manager_module.DATA_DIR = original_data_dir

            self.assertEqual(record["sheet_name"], "Funded Invoices")
            self.assertEqual(record["source_profile"], "extraction_file")
            self.assertTrue(record["available_cohorts"])
            self.assertTrue(all(item["date"].startswith("2025-01-28") for item in record["available_cohorts"]))
            self.assertEqual(len(record["available_cohorts"]), 1)

            process_dir = Path(record["process_dir"])
            self.assertTrue((process_dir / "extraction.xlsx").exists())


if __name__ == "__main__":
    unittest.main()
