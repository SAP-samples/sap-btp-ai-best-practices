"""
Excel Generator Service

Generates summary Excel files with Funded and Non-Funded invoice sheets.
"""

import logging
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from ...config.eligibility_config import get_output_directory
from ...models.eligibility import FundedInvoice, NonFundedInvoice

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Generator for creating summary Excel files with eligibility results.

    Creates workbooks with two sheets:
    - Funded Invoices: Invoices that passed all eligibility rules
    - Non-Funded Invoices: Invoices that failed one or more rules (with rejection reasons)
    """

    # Column configurations for each sheet
    FUNDED_COLUMNS = [
        ("PROGRAMA", "programa"),
        ("ID SELLER", "seller_id"),
        ("SELLER", "seller_name"),
        ("ID DEBTOR", "debtor_id"),
        ("DEBTOR", "debtor_name"),
        ("INVOICE REF", "invoice_ref"),
        ("ORIGINAL CURRENCY", "original_currency"),
        ("ISSUANCE DATE", "issuance_date"),
        ("DUE DATE", "due_date"),
        ("AMOUNT ORIGINAL", "amount_original"),
        ("AMOUNT EUR", "amount_eur"),
        ("PURCHASE DATE", "purchase_date"),
        ("DAYS TO DUE", "days_to_due"),
        ("TENOR", "tenor"),
    ]

    NON_FUNDED_COLUMNS = [
        ("PROGRAMA", "programa"),
        ("ID SELLER", "seller_id"),
        ("SELLER", "seller_name"),
        ("ID DEBTOR", "debtor_id"),
        ("DEBTOR", "debtor_name"),
        ("INVOICE REF", "invoice_ref"),
        ("ORIGINAL CURRENCY", "original_currency"),
        ("ISSUANCE DATE", "issuance_date"),
        ("DUE DATE", "due_date"),
        ("AMOUNT ORIGINAL", "amount_original"),
        ("AMOUNT EUR", "amount_eur"),
        ("PURCHASE DATE", "purchase_date"),
        ("REJECTION REASON", "rejection_reason"),
        ("FAILED RULES", "failed_rules_str"),
    ]

    # Styling constants
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    FUNDED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    NON_FUNDED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self, output_dir: Optional[Path] = None):
        """
        Initialize the Excel generator.

        Args:
            output_dir: Directory for saving generated files (defaults to config path)
        """
        self.output_dir = output_dir or get_output_directory()
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _funded_to_dict(self, invoice: FundedInvoice) -> dict:
        """Convert a FundedInvoice to a dictionary for DataFrame."""
        return {
            "programa": invoice.programa,
            "seller_id": invoice.seller_id,
            "seller_name": invoice.seller_name,
            "debtor_id": invoice.debtor_id,
            "debtor_name": invoice.debtor_name,
            "invoice_ref": invoice.invoice_ref,
            "original_currency": invoice.original_currency,
            "issuance_date": invoice.issuance_date,
            "due_date": invoice.due_date,
            "amount_original": float(invoice.amount_original) if invoice.amount_original else None,
            "amount_eur": float(invoice.amount_eur) if invoice.amount_eur else None,
            "purchase_date": invoice.purchase_date,
            "days_to_due": invoice.days_to_due,
            "tenor": invoice.tenor,
        }

    def _non_funded_to_dict(self, invoice: NonFundedInvoice) -> dict:
        """Convert a NonFundedInvoice to a dictionary for DataFrame."""
        return {
            "programa": invoice.programa,
            "seller_id": invoice.seller_id,
            "seller_name": invoice.seller_name,
            "debtor_id": invoice.debtor_id,
            "debtor_name": invoice.debtor_name,
            "invoice_ref": invoice.invoice_ref,
            "original_currency": invoice.original_currency,
            "issuance_date": invoice.issuance_date,
            "due_date": invoice.due_date,
            "amount_original": float(invoice.amount_original) if invoice.amount_original else None,
            "amount_eur": float(invoice.amount_eur) if invoice.amount_eur else None,
            "purchase_date": invoice.purchase_date,
            "rejection_reason": invoice.rejection_reason,
            "failed_rules_str": ", ".join(invoice.failed_rules),
        }

    def _style_header_row(self, ws, num_columns: int) -> None:
        """Apply styling to the header row."""
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.BORDER

    def _style_data_rows(
        self,
        ws,
        num_rows: int,
        num_columns: int,
        fill: Optional[PatternFill] = None,
    ) -> None:
        """Apply styling to data rows."""
        for row in range(2, num_rows + 2):  # Start from row 2 (after header)
            for col in range(1, num_columns + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.BORDER
                cell.alignment = Alignment(vertical="center")
                if fill:
                    cell.fill = fill

    def _auto_adjust_columns(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            # Set width with some padding, max 50
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def generate_summary(
        self,
        funded: List[FundedInvoice],
        non_funded: List[NonFundedInvoice],
        filename: Optional[str] = None,
    ) -> bytes:
        """
        Generate a summary Excel file with Funded and Non-Funded sheets.

        Args:
            funded: List of funded invoices
            non_funded: List of non-funded invoices
            filename: Optional filename to save (if not provided, only returns bytes)

        Returns:
            Excel file content as bytes
        """
        logger.info(
            f"Generating summary Excel: {len(funded)} funded, {len(non_funded)} non-funded"
        )

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Create Funded Invoices sheet
        ws_funded = wb.create_sheet("Funded Invoices")
        self._write_funded_sheet(ws_funded, funded)

        # Create Non-Funded Invoices sheet
        ws_non_funded = wb.create_sheet("Non-Funded Invoices")
        self._write_non_funded_sheet(ws_non_funded, non_funded)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        content = output.getvalue()

        # Optionally save to file
        if filename:
            filepath = self.output_dir / filename
            with open(filepath, "wb") as f:
                f.write(content)
            logger.info(f"Saved summary Excel to {filepath}")

        return content

    def _write_funded_sheet(self, ws, funded: List[FundedInvoice]) -> None:
        """Write the Funded Invoices sheet."""
        # Write headers
        for col, (header, _) in enumerate(self.FUNDED_COLUMNS, 1):
            ws.cell(row=1, column=col, value=header)

        # Convert to DataFrame for easy writing
        if funded:
            data = [self._funded_to_dict(inv) for inv in funded]
            df = pd.DataFrame(data)

            # Write data rows
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, (_, field) in enumerate(self.FUNDED_COLUMNS, 1):
                    value = row_data.get(field)
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Apply styling
        num_columns = len(self.FUNDED_COLUMNS)
        self._style_header_row(ws, num_columns)
        self._style_data_rows(ws, len(funded), num_columns, self.FUNDED_FILL)
        self._auto_adjust_columns(ws)

        # Freeze header row
        ws.freeze_panes = "A2"

    def _write_non_funded_sheet(self, ws, non_funded: List[NonFundedInvoice]) -> None:
        """Write the Non-Funded Invoices sheet."""
        # Write headers
        for col, (header, _) in enumerate(self.NON_FUNDED_COLUMNS, 1):
            ws.cell(row=1, column=col, value=header)

        # Write data rows
        if non_funded:
            data = [self._non_funded_to_dict(inv) for inv in non_funded]

            for row_idx, row_data in enumerate(data, 2):
                for col_idx, (_, field) in enumerate(self.NON_FUNDED_COLUMNS, 1):
                    value = row_data.get(field)
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Apply styling
        num_columns = len(self.NON_FUNDED_COLUMNS)
        self._style_header_row(ws, num_columns)
        self._style_data_rows(ws, len(non_funded), num_columns, self.NON_FUNDED_FILL)
        self._auto_adjust_columns(ws)

        # Freeze header row
        ws.freeze_panes = "A2"

    def get_generated_file_path(self, filename: str) -> Optional[Path]:
        """
        Get the full path to a generated file.

        Args:
            filename: The filename to check

        Returns:
            Path if file exists, None otherwise
        """
        filepath = self.output_dir / filename
        if filepath.exists():
            return filepath
        return None

    def generate_filename(self, prefix: str = "eligibility_summary") -> str:
        """
        Generate a unique filename with timestamp.

        Args:
            prefix: Filename prefix

        Returns:
            Filename with timestamp
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.xlsx"
