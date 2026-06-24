from __future__ import annotations

import base64
import csv
from io import BytesIO, StringIO
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from .models import EmailAttachment


AttachmentSource = Literal["gmail", "manual"]

MAX_ATTACHMENTS_PER_EMAIL = 5
MAX_ATTACHMENT_BYTES = 10 * 1024 * 1024
MAX_TABULAR_ROWS = 120
MAX_TABULAR_COLUMNS = 40
MAX_TABULAR_TEXT_CHARS = 25_000

SUPPORTED_EXTENSION_MIME_TYPES = {
    "pdf": "application/pdf",
    "csv": "text/csv",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


class UnsupportedAttachmentTypeError(ValueError):
    """Raised when an uploaded or fetched attachment type is not supported."""


class AttachmentTooLargeError(ValueError):
    """Raised when an attachment exceeds the configured byte limit."""


def sanitize_filename(filename: str) -> str:
    """Return a filesystem-neutral display filename.

    Args:
        filename: Raw filename from Gmail or a browser upload.

    Returns:
        Sanitized filename without path segments or NUL characters.
    """
    sanitized = Path(filename.replace("\x00", "")).name.strip()
    return sanitized or "attachment"


def normalized_file_extension(filename: str) -> str:
    """Return the lowercase supported extension without the leading dot.

    Args:
        filename: Sanitized or raw attachment filename.

    Returns:
        Lowercase extension without a leading dot.
    """
    return Path(filename).suffix.lower().removeprefix(".")


def normalized_mime_type(filename: str, mime_type: str | None) -> str:
    """Normalize MIME type for supported attachment files.

    Args:
        filename: Attachment filename used for extension inference.
        mime_type: Raw MIME type from Gmail or browser upload.

    Returns:
        MIME type suitable for persisted metadata and downloads.
    """
    extension = normalized_file_extension(filename)
    raw_mime = (mime_type or "").split(";", 1)[0].strip().lower()
    if extension in SUPPORTED_EXTENSION_MIME_TYPES:
        if not raw_mime or raw_mime in {"application/octet-stream", "application/vnd.ms-excel"}:
            return SUPPORTED_EXTENSION_MIME_TYPES[extension]
        return raw_mime
    return raw_mime or "application/octet-stream"


def ensure_supported_attachment(filename: str, mime_type: str | None) -> tuple[str, str]:
    """Validate an attachment type and return normalized extension/MIME type.

    Args:
        filename: Attachment filename.
        mime_type: Attachment MIME type.

    Returns:
        Tuple of supported extension and normalized MIME type.

    Raises:
        UnsupportedAttachmentTypeError: If the extension is not PDF, CSV, or XLSX.
    """
    extension = normalized_file_extension(filename)
    if extension not in SUPPORTED_EXTENSION_MIME_TYPES:
        raise UnsupportedAttachmentTypeError(
            f"Unsupported attachment type for {filename}. Supported types are PDF, CSV, and XLSX."
        )
    return extension, normalized_mime_type(filename, mime_type)


def enforce_attachment_size(content: bytes, max_size_bytes: int = MAX_ATTACHMENT_BYTES) -> None:
    """Reject attachment bytes that exceed the configured size limit.

    Args:
        content: Raw attachment bytes.
        max_size_bytes: Maximum allowed byte size.

    Raises:
        AttachmentTooLargeError: If content is larger than max_size_bytes.
    """
    if len(content) > max_size_bytes:
        raise AttachmentTooLargeError(
            f"Attachment is {len(content)} bytes, exceeding the {max_size_bytes} byte limit."
        )


def decode_base64url_bytes(data: str) -> bytes:
    """Decode Gmail base64url attachment data.

    Args:
        data: Base64url text without guaranteed padding.

    Returns:
        Decoded bytes.
    """
    padded = data + "=" * (-len(data) % 4)
    return base64.urlsafe_b64decode(padded.encode("utf-8"))


def bounded_text(value: str, max_chars: int = MAX_TABULAR_TEXT_CHARS) -> str:
    """Trim long model attachment text while marking truncation.

    Args:
        value: Raw text.
        max_chars: Maximum characters to retain.

    Returns:
        Bounded text value.
    """
    if len(value) <= max_chars:
        return value
    return f"{value[:max_chars]}\n[Attachment text truncated after {max_chars} characters.]"


def csv_bytes_to_text(content: bytes) -> str:
    """Convert CSV bytes into a bounded, normalized CSV text preview.

    Args:
        content: Raw CSV bytes.

    Returns:
        CSV text suitable for prompt context.
    """
    decoded = content.decode("utf-8-sig", errors="replace")
    input_buffer = StringIO(decoded)
    output_buffer = StringIO()
    writer = csv.writer(output_buffer, lineterminator="\n")
    for index, row in enumerate(csv.reader(input_buffer)):
        if index >= MAX_TABULAR_ROWS:
            writer.writerow([f"[CSV preview truncated after {MAX_TABULAR_ROWS} rows]"])
            break
        writer.writerow(row[:MAX_TABULAR_COLUMNS])
    return bounded_text(output_buffer.getvalue().strip())


def xlsx_bytes_to_text(content: bytes) -> str:
    """Convert XLSX bytes into bounded sheet-by-sheet CSV-like text.

    Args:
        content: Raw XLSX bytes.

    Returns:
        CSV-like sheet previews suitable for prompt context.
    """
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sections: list[str] = []
    try:
        for sheet in workbook.worksheets:
            output_buffer = StringIO()
            writer = csv.writer(output_buffer, lineterminator="\n")
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_index >= MAX_TABULAR_ROWS:
                    writer.writerow([f"[XLSX preview truncated after {MAX_TABULAR_ROWS} rows]"])
                    break
                writer.writerow(["" if value is None else value for value in row[:MAX_TABULAR_COLUMNS]])
            sections.append(f"Sheet: {sheet.title}\n{output_buffer.getvalue().strip()}")
    finally:
        workbook.close()
    return bounded_text("\n\n".join(sections).strip())


def tabular_attachment_text(filename: str, extension: str, content: bytes) -> str | None:
    """Return bounded text for tabular attachments.

    Args:
        filename: Sanitized attachment filename.
        extension: Supported extension without a leading dot.
        content: Raw attachment bytes.

    Returns:
        Text preview for CSV/XLSX, or None for PDFs.
    """
    if extension == "csv":
        return csv_bytes_to_text(content)
    if extension == "xlsx":
        return xlsx_bytes_to_text(content)
    _ = filename
    return None


def build_email_attachment(
    filename: str,
    mime_type: str | None,
    content: bytes,
    source: AttachmentSource,
    provider_attachment_id: str | None = None,
    gmail_message_id: str | None = None,
    max_size_bytes: int = MAX_ATTACHMENT_BYTES,
) -> EmailAttachment:
    """Build a validated attachment model from raw bytes.

    Args:
        filename: Raw attachment filename.
        mime_type: Raw MIME type.
        content: Raw attachment bytes.
        source: Input route that provided the attachment.
        provider_attachment_id: Gmail attachment id when available.
        gmail_message_id: Owning Gmail or synthetic manual message id.
        max_size_bytes: Maximum allowed byte size.

    Returns:
        EmailAttachment ready for HANA persistence.
    """
    safe_filename = sanitize_filename(filename)
    extension, normalized_mime = ensure_supported_attachment(safe_filename, mime_type)
    enforce_attachment_size(content, max_size_bytes=max_size_bytes)
    return EmailAttachment(
        gmail_message_id=gmail_message_id,
        source=source,
        provider_attachment_id=provider_attachment_id,
        filename=safe_filename,
        mime_type=normalized_mime,
        file_extension=extension,  # type: ignore[arg-type]
        size_bytes=len(content),
        content_base64=base64.b64encode(content).decode("ascii"),
        text_preview=tabular_attachment_text(safe_filename, extension, content),
    )


def attachment_bytes(attachment: EmailAttachment) -> bytes:
    """Decode raw attachment bytes from a persisted model.

    Args:
        attachment: Persisted attachment model with base64 content.

    Returns:
        Raw attachment bytes.
    """
    return base64.b64decode(attachment.content_base64.encode("ascii"))


def attachment_to_responses_content_blocks(attachment: EmailAttachment) -> list[dict[str, str]]:
    """Build OpenAI Responses content blocks for one attachment.

    Args:
        attachment: Attachment model.

    Returns:
        Responses API content blocks for PDF or tabular text inputs.
    """
    if attachment.file_extension == "pdf":
        return [
            {
                "type": "input_file",
                "filename": attachment.filename,
                "file_data": f"data:application/pdf;base64,{attachment.content_base64}",
            }
        ]
    return [
        {
            "type": "input_text",
            "text": (
                f"Attachment: {attachment.filename} ({attachment.mime_type}, {attachment.size_bytes} bytes)\n"
                f"{attachment.text_preview or ''}"
            ),
        }
    ]


def attachments_to_responses_content_blocks(attachments: list[EmailAttachment]) -> list[dict[str, str]]:
    """Build Responses content blocks for all supported attachments.

    Args:
        attachments: Attachments to include in model context.

    Returns:
        Flat list of Responses API content blocks.
    """
    blocks: list[dict[str, str]] = []
    for attachment in attachments:
        blocks.extend(attachment_to_responses_content_blocks(attachment))
    return blocks
