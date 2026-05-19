from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path

import fitz
import openpyxl


PAGE_WIDTH = 595
PAGE_HEIGHT = 842
MARGIN_X = 42
MARGIN_Y = 40

BLACK = (0.08, 0.08, 0.08)
MID_GREY = (0.45, 0.45, 0.45)
LIGHT_GREY = (0.9, 0.92, 0.94)
BLUE = (0.40, 0.67, 0.91)
DARK_BLUE = (0.13, 0.22, 0.39)
RED = (0.82, 0.12, 0.15)


@dataclass
class Commodity:
    code: str
    high_level: str
    low_level: str
    keywords: str
    spend_type: str


@dataclass
class SupplierAssignment:
    supplier_name: str
    business_partner_id: str
    material_group: str


def data_dir() -> Path:
    return Path(__file__).resolve().parents[2] / "generated_reference_data"


def output_dir() -> Path:
    return Path(__file__).resolve().parents[1] / "outputs" / "api"


def load_commodity_map() -> dict[str, Commodity]:
    workbook = openpyxl.load_workbook(
        data_dir() / "commodity_catalog.synthetic.xlsx",
        data_only=True,
    )
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    header = rows[0]
    commodity_map: dict[str, Commodity] = {}
    for row in rows[1:]:
        record = dict(zip(header, row))
        commodity_map[record["CODE"]] = Commodity(
            code=record["CODE"],
            high_level=record["DESCRIPTION_HIGH_LEVEL"],
            low_level=record["DESCRIPTION_LOW_LEVEL"],
            keywords=record["DETAILED_DESCRIPTION_KEYWORDS"],
            spend_type=record["SPEND_TYPE"],
        )
    return commodity_map


def load_suppliers() -> list[SupplierAssignment]:
    with (data_dir() / "supplier_material_groups.synthetic.csv").open(newline="") as handle:
        reader = csv.DictReader(handle)
        return [
            SupplierAssignment(
                supplier_name=row["SUPPLIER_NAME"],
                business_partner_id=row["BUSINESS_PARTNER_ID"],
                material_group=row["MATERIAL_GROUP"],
            )
            for row in reader
        ]


def supplier_for_code(rows: list[SupplierAssignment], code: str, preferred: str = "") -> SupplierAssignment:
    matches = [row for row in rows if row.material_group == code]
    if preferred:
        for row in matches:
            if preferred.lower() in row.supplier_name.lower():
                return row
    if not matches:
        raise ValueError(f"No supplier rows found for material group {code}")
    return matches[0]


def draw_textbox(
    page: fitz.Page,
    rect: fitz.Rect,
    text: str,
    *,
    fontsize: float = 11,
    fontname: str = "helv",
    color: tuple[float, float, float] = BLACK,
    align: int = 0,
    lineheight: float | None = None,
) -> None:
    page.insert_textbox(
        rect,
        text,
        fontsize=fontsize,
        fontname=fontname,
        color=color,
        align=align,
        lineheight=lineheight or 1.2,
    )


def draw_label_value(page: fitz.Page, x: float, y: float, label: str, value: str, width: float = 205) -> None:
    draw_textbox(
        page,
        fitz.Rect(x, y, x + 80, y + 18),
        label,
        fontsize=10,
        fontname="hebo",
    )
    draw_textbox(
        page,
        fitz.Rect(x + 82, y, x + width, y + 18),
        value,
        fontsize=10,
    )


def draw_meta_block(page: fitz.Page, x: float, y: float, label: str, value: str, width: float = 120) -> None:
    draw_textbox(
        page,
        fitz.Rect(x, y, x + width, y + 18),
        label,
        fontsize=8.6,
        fontname="hebo",
        color=MID_GREY,
    )
    draw_textbox(
        page,
        fitz.Rect(x, y + 15, x + width, y + 36),
        value,
        fontsize=10,
        fontname="hebo",
        color=BLACK,
    )


def draw_footer(page: fitz.Page, left: str, center: str, right: str) -> None:
    y = PAGE_HEIGHT - 46
    page.draw_line((MARGIN_X, y - 8), (PAGE_WIDTH - MARGIN_X, y - 8), color=LIGHT_GREY, width=0.8)
    draw_textbox(page, fitz.Rect(MARGIN_X, y, 210, y + 28), left, fontsize=8, color=MID_GREY)
    draw_textbox(
        page,
        fitz.Rect(210, y, 390, y + 28),
        center,
        fontsize=8,
        color=MID_GREY,
        align=1,
    )
    draw_textbox(
        page,
        fitz.Rect(390, y, PAGE_WIDTH - MARGIN_X, y + 28),
        right,
        fontsize=8,
        color=MID_GREY,
        align=2,
    )


def draw_table(
    page: fitz.Page,
    top_y: float,
    headers: list[str],
    rows: list[list[str]],
    widths: list[float],
    aligns: list[int],
    row_height: float = 36,
    font_size: float = 9.5,
) -> float:
    x_positions = [MARGIN_X]
    for width in widths[:-1]:
        x_positions.append(x_positions[-1] + width)

    header_bottom = top_y + 18
    page.draw_line((MARGIN_X, header_bottom), (MARGIN_X + sum(widths), header_bottom), color=BLACK, width=0.8)
    for idx, header in enumerate(headers):
        cell = fitz.Rect(
            x_positions[idx] + 2,
            top_y - 2,
            x_positions[idx] + widths[idx] - 4,
            header_bottom - 1,
        )
        draw_textbox(page, cell, header, fontsize=9, fontname="hebo", align=aligns[idx])

    y = header_bottom + 2
    for row in rows:
        page.draw_line((MARGIN_X, y + row_height), (MARGIN_X + sum(widths), y + row_height), color=LIGHT_GREY, width=0.8)
        for idx, value in enumerate(row):
            cell = fitz.Rect(
                x_positions[idx] + 2,
                y + 2,
                x_positions[idx] + widths[idx] - 4,
                y + row_height - 2,
            )
            draw_textbox(page, cell, value, fontsize=font_size, align=aligns[idx], lineheight=1.15)
        y += row_height
    return y


def add_document_metadata(doc: fitz.Document, title: str, author: str) -> None:
    doc.set_metadata(
        {
            "title": title,
            "author": author,
            "subject": "Synthetic anonymized PDF for commodity code matching tests",
            "creator": "generate_sample_pdfs.py",
            "producer": "PyMuPDF",
            "keywords": "synthetic, anonymized, commodity code, test pdf",
        }
    )


def create_transport_offer(doc: fitz.Document, supplier: SupplierAssignment, commodity: Commodity) -> None:
    page = doc.new_page(width=PAGE_WIDTH, height=PAGE_HEIGHT)

    draw_textbox(page, fitz.Rect(28, 34, 200, 74), "GRANITE", fontsize=24, fontname="hebo", color=RED)
    draw_textbox(page, fitz.Rect(28, 58, 210, 76), "digital logistics services", fontsize=9, color=MID_GREY)

    draw_textbox(page, fitz.Rect(270, 34, 430, 70), "Offert", fontsize=22, fontname="hebo")
    draw_textbox(page, fitz.Rect(270, 62, 430, 84), "OFF-A9001", fontsize=12, fontname="hebo")
    draw_textbox(page, fitz.Rect(520, 45, 560, 64), "Sida 1", fontsize=9, align=2)

    draw_textbox(
        page,
        fitz.Rect(MARGIN_X, 92, 235, 165),
        "Leveransadress\nNorthwind Motion AB\nPrototype Logistics Team\nSE-417 55 Gothenburg",
        fontsize=10,
        fontname="hebo",
        lineheight=1.25,
    )
    draw_textbox(
        page,
        fitz.Rect(325, 92, 555, 175),
        "Faktureringsadress\nAurora Drive Manufacturing AB\nAccounts Payable - Unit 48\nSE-415 31 Gothenburg",
        fontsize=10,
        fontname="hebo",
        lineheight=1.25,
    )

    draw_label_value(page, MARGIN_X, 204, "Leveransvillkor:", "Fraktfritt inom region")
    draw_label_value(page, MARGIN_X, 222, "Betalningsvillkor:", "45 dagar netto")
    draw_label_value(page, MARGIN_X, 240, "Er kontaktperson", "Hanna Lind")
    draw_label_value(page, MARGIN_X, 258, "Saljare:", "Mikael Noren")
    draw_label_value(page, MARGIN_X, 276, "E-mail:", "offers@granite-digital.example")
    draw_label_value(page, MARGIN_X, 294, "Telefon:", "+46 31 700 11 40")

    draw_label_value(page, 325, 204, "Datum:", "2026-04-01")
    draw_label_value(page, 325, 222, "Kundnr:", "C-50392")
    draw_label_value(page, 325, 258, "Giltig t o m", "2026-04-22")

    headers = ["Nr", "Benamning", "Antal", "Enhet", "A-pris", "Rabatt\n%", "Belopp"]
    widths = [44, 220, 40, 42, 62, 38, 65]
    aligns = [0, 0, 2, 0, 2, 2, 2]
    rows = [
        ["74010", "Regional shuttle transport between battery lab and assembly site", "24", "Trip", "1 850,00", "0", "44 400,00"],
        ["74022", "Dedicated fleet movement for validation vehicles and support vans", "12", "Trip", "4 200,00", "0", "50 400,00"],
        ["74035", "Rail-linked parts transfer and handling coordination for recurring deliveries", "8", "Lot", "6 100,00", "0", "48 800,00"],
    ]
    bottom_y = draw_table(page, 328, headers, rows, widths, aligns, row_height=42)

    draw_textbox(
        page,
        fitz.Rect(302, bottom_y + 14, 470, bottom_y + 34),
        "Total exkl. moms SEK",
        fontsize=10,
        fontname="hebo",
        align=2,
    )
    draw_textbox(
        page,
        fitz.Rect(472, bottom_y + 14, 555, bottom_y + 34),
        "143 600,00",
        fontsize=10,
        fontname="hebo",
        align=2,
    )

    note = (
        f"Supplier: {supplier.supplier_name} ({supplier.business_partner_id})\n"
        f"Typical usage in this quote matches the synthetic reference family "
        f"'{commodity.low_level}'."
    )
    draw_textbox(page, fitz.Rect(MARGIN_X, bottom_y + 56, 555, bottom_y + 104), note, fontsize=9, color=MID_GREY)

    draw_footer(
        page,
        "Granite Digital Partners 017\nSynthetic supplier sample",
        "BP-000004\nTransport quote test document",
        "org no. 556900-1240\nwww.granite-digital.example",
    )


def create_lab_quotation(doc: fitz.Document, supplier: SupplierAssignment, commodity: Commodity) -> None:
    page1 = doc.new_page(width=PAGE_WIDTH, height=PAGE_HEIGHT)
    draw_textbox(page1, fitz.Rect(MARGIN_X, 26, 220, 70), "JUNIPER", fontsize=26, fontname="hebo", color=RED)
    draw_textbox(page1, fitz.Rect(MARGIN_X, 57, 220, 74), "industrial validation systems", fontsize=9, color=MID_GREY)
    draw_meta_block(page1, 304, 28, "ISSUED BY", "Linnea Forsberg", width=120)
    draw_meta_block(page1, 304, 56, "DOC. NAME", "Quotation_Lab_Validation_v1", width=150)
    draw_meta_block(page1, 470, 28, "DATE", "2026-04-01", width=84)
    draw_meta_block(page1, 470, 56, "PAGE", "1 (2)", width=84)

    footer = (
        f"{supplier.supplier_name}\n"
        "Box 2104\n"
        "SE-461 30 Trollhattan, Sweden"
    )
    draw_footer(
        page1,
        footer,
        "TEL +46 (0)520 44 18 00\ninfo@juniper-industrial.example",
        f"BP {supplier.business_partner_id}\nSynthetic quotation sample",
    )

    draw_textbox(page1, fitz.Rect(350, 110, 545, 182), "Aurora Motion Systems BV\nAttn. Sofia Ahlgren\nRidge Park 12\n2018 Antwerp", fontsize=10, fontname="hebo")
    draw_textbox(page1, fitz.Rect(MARGIN_X, 128, 360, 164), "Quotation No: LAB-2401 v1", fontsize=18, fontname="hebo")
    draw_textbox(page1, fitz.Rect(MARGIN_X, 156, 545, 182), "Thank you for your inquiry. We are pleased to quote a modular validation and calibration package for recurring lab testing.", fontsize=10.5)

    sections = [
        (
            "1   Purpose of quote",
            "The purpose of this quote is to provide measurement and validation equipment for repeated performance testing in an anonymized pilot environment.",
        ),
        (
            "2   Description",
            "The proposed package contains a configurable test bench, a calibration sensor pack, and a controlled commissioning activity. The scope is aligned with quality and performance testing work where traceable measurement data is required.",
        ),
        (
            "3   Included equipment",
            "The package supports:\n- modular test bench fixtures\n- calibration and measurement devices\n- reporting template for laboratory acceptance\n- operator handover for validation staff",
        ),
        (
            "4   Installation/commissioning",
            "A functional acceptance check takes place at the supplier site before delivery. Final commissioning is completed at the customer lab within one working day.",
        ),
        (
            "5   Prerequisites and limitations",
            "Customer supplies the target samples, electrical drop, and local safety approval. Quote excludes civil works and permanent building modifications.",
        ),
    ]
    y = 196
    for heading, body in sections:
        draw_textbox(page1, fitz.Rect(MARGIN_X, y, 545, y + 24), heading, fontsize=12, fontname="hebo")
        y += 22
        draw_textbox(page1, fitz.Rect(MARGIN_X + 20, y, 545, y + 68), body, fontsize=10.3, lineheight=1.22)
        y += 76

    page2 = doc.new_page(width=PAGE_WIDTH, height=PAGE_HEIGHT)
    draw_textbox(page2, fitz.Rect(MARGIN_X, 26, 220, 70), "JUNIPER", fontsize=26, fontname="hebo", color=RED)
    draw_textbox(page2, fitz.Rect(MARGIN_X, 57, 220, 74), "industrial validation systems", fontsize=9, color=MID_GREY)
    draw_meta_block(page2, 304, 28, "ISSUED BY", "Linnea Forsberg", width=120)
    draw_meta_block(page2, 304, 56, "DOC. NAME", "Quotation_Lab_Validation_v1", width=150)
    draw_meta_block(page2, 470, 28, "DATE", "2026-04-01", width=84)
    draw_meta_block(page2, 470, 56, "PAGE", "2 (2)", width=84)
    draw_footer(
        page2,
        footer,
        "TEL +46 (0)520 44 18 00\ninfo@juniper-industrial.example",
        f"BP {supplier.business_partner_id}\nSynthetic quotation sample",
    )

    draw_textbox(page2, fitz.Rect(MARGIN_X, 128, 545, 152), "6   Delivery", fontsize=12, fontname="hebo")
    draw_textbox(page2, fitz.Rect(MARGIN_X + 20, 152, 545, 192), "Delivery time 8-10 working weeks from written order. Delivery terms DAP Antwerp including standard packaging. Warranty 12 months.", fontsize=10.3)

    draw_textbox(page2, fitz.Rect(MARGIN_X, 210, 545, 236), "7   Price exclusive of VAT, SEK", fontsize=12, fontname="hebo")
    headers = ["Description", "Qty", "Unit price", "Total"]
    widths = [328, 48, 82, 95]
    aligns = [0, 2, 2, 2]
    rows = [
        ["Modular validation rig for laboratory performance testing", "1", "185 000", "185 000"],
        ["Calibration sensor pack with traceable measurement devices", "1", "54 000", "54 000"],
        ["Commissioning and reporting setup for test bench acceptance", "1", "32 500", "32 500"],
    ]
    bottom_y = draw_table(page2, 236, headers, rows, widths, aligns, row_height=40, font_size=10)
    draw_textbox(page2, fitz.Rect(355, bottom_y + 10, 450, bottom_y + 30), "Total SEK", fontsize=10, fontname="hebo", align=2)
    draw_textbox(page2, fitz.Rect(450, bottom_y + 10, 545, bottom_y + 30), "271 500", fontsize=10, fontname="hebo", align=2)

    draw_textbox(page2, fitz.Rect(MARGIN_X, bottom_y + 46, 545, bottom_y + 96), "8   Validity of the quotation", fontsize=12, fontname="hebo")
    draw_textbox(page2, fitz.Rect(MARGIN_X + 20, bottom_y + 70, 545, bottom_y + 118), "This offer is valid until 2026-05-02. Synthetic reference family used for the content design: measurement, validation, calibration, and laboratory equipment.", fontsize=10.3)


def create_digital_cover_quote(doc: fitz.Document, supplier: SupplierAssignment, cloud: Commodity, security: Commodity) -> None:
    page1 = doc.new_page(width=PAGE_WIDTH, height=PAGE_HEIGHT)

    page1.draw_rect(fitz.Rect(0, 120, PAGE_WIDTH, 310), fill=BLUE, color=BLUE)
    page1.draw_rect(fitz.Rect(365, 120, PAGE_WIDTH, 310), fill=(0.92, 0.94, 0.97), color=(0.92, 0.94, 0.97))
    page1.draw_rect(fitz.Rect(140, 270, 460, 610), fill=(0.95, 0.96, 0.98), color=(0.76, 0.8, 0.86), width=1)

    for idx in range(6):
        x = 175 + idx * 42
        page1.draw_line((x, 320), (x + 22, 560), color=(0.72, 0.78, 0.86), width=1)
    for idx in range(5):
        y = 350 + idx * 45
        page1.draw_line((170, y), (430, y), color=(0.72, 0.78, 0.86), width=1)

    draw_textbox(page1, fitz.Rect(34, 24, 280, 92), "CREST", fontsize=34, fontname="hebo", color=BLACK)
    draw_textbox(page1, fitz.Rect(36, 78, 230, 98), "Quotation date: 1 April 2026", fontsize=10, color=DARK_BLUE)
    draw_textbox(page1, fitz.Rect(400, 40, 560, 88), "Northwind Assembly Europe NV\nHarbor Avenue 18\n2000 Antwerp", fontsize=10, fontname="hebo", align=2, color=DARK_BLUE)

    draw_textbox(page1, fitz.Rect(58, 150, 330, 195), "QUOTATION 8803", fontsize=23, fontname="hebo", color=DARK_BLUE)
    draw_textbox(page1, fitz.Rect(58, 198, 330, 248), "Shared digital workload hosting\nand security operations", fontsize=16, color=DARK_BLUE)
    draw_textbox(page1, fitz.Rect(58, 640, 260, 720), "Customer Contact Details\nNorthwind Assembly Europe NV\nHarbor Avenue 18\n2000 Antwerp", fontsize=11, fontname="hebo", color=DARK_BLUE)
    draw_textbox(page1, fitz.Rect(360, 640, 560, 720), "Issued by\nCrest Network Partners 014\nBusiness Partner ID: BP-001602", fontsize=11, fontname="hebo", align=2, color=DARK_BLUE)
    draw_footer(page1, "Synthetic branded quotation sample", "Managed services and hosting", "Page 1")

    page2 = doc.new_page(width=PAGE_WIDTH, height=PAGE_HEIGHT)
    draw_textbox(page2, fitz.Rect(MARGIN_X, 40, 420, 74), "Quotation 8803 - service summary", fontsize=19, fontname="hebo")
    draw_textbox(page2, fitz.Rect(400, 44, 555, 74), "Page 2", fontsize=10, fontname="hebo", align=2)
    draw_textbox(page2, fitz.Rect(MARGIN_X, 82, 545, 122), "The proposal covers a managed cloud platform landing zone, shared hosting, and the ongoing security operations needed to support anonymized testing environments.", fontsize=10.5)

    headers = ["Line", "Service description", "Months", "Monthly fee", "Amount"]
    widths = [44, 280, 54, 82, 93]
    aligns = [0, 0, 2, 2, 2]
    rows = [
        ["10", "Managed cloud landing zone and shared hosting service for core workloads", "12", "14 800", "177 600"],
        ["20", "Identity controls, security monitoring, and digital operations support", "12", "9 900", "118 800"],
        ["30", "Environment setup, runbook handover, and service transition workshop", "1", "24 500", "24 500"],
    ]
    bottom_y = draw_table(page2, 140, headers, rows, widths, aligns, row_height=44, font_size=10)

    draw_textbox(page2, fitz.Rect(342, bottom_y + 10, 450, bottom_y + 28), "Total SEK", fontsize=10, fontname="hebo", align=2)
    draw_textbox(page2, fitz.Rect(450, bottom_y + 10, 545, bottom_y + 28), "320 900", fontsize=10, fontname="hebo", align=2)

    draw_textbox(page2, fitz.Rect(MARGIN_X, bottom_y + 52, 545, bottom_y + 72), "Reference families used in the synthetic content", fontsize=11, fontname="hebo")
    body = (
        f"{cloud.high_level}: {cloud.low_level}.\n"
        f"Keywords: {cloud.keywords}.\n\n"
        f"{security.high_level}: {security.low_level}.\n"
        f"Keywords: {security.keywords}."
    )
    draw_textbox(page2, fitz.Rect(MARGIN_X, bottom_y + 78, 545, bottom_y + 176), body, fontsize=10.2, lineheight=1.2)
    draw_textbox(page2, fitz.Rect(MARGIN_X, bottom_y + 196, 545, bottom_y + 258), "Commercial terms: payment 30 days net, service start within 15 working days after order, validity until 2026-05-10.", fontsize=10.2)
    draw_footer(page2, supplier.supplier_name, supplier.business_partner_id, "Synthetic quotation sample")


def save_document(path: Path, build_fn) -> None:
    document = fitz.open()
    build_fn(document)
    document.save(path, garbage=4, deflate=True)
    document.close()


def main() -> None:
    commodities = load_commodity_map()
    suppliers = load_suppliers()
    destination = output_dir()
    destination.mkdir(parents=True, exist_ok=True)

    transport_supplier = supplier_for_code(suppliers, "RC1401", preferred="Granite")
    lab_supplier = supplier_for_code(suppliers, "RC1202", preferred="Juniper")
    digital_supplier = supplier_for_code(suppliers, "RC1101", preferred="Crest")

    transport_path = destination / "OFF-A9001-anonymized.pdf"
    lab_path = destination / "Quotation-LAB-2401-anonymized.pdf"
    digital_path = destination / "Quotation-8803-anonymized.pdf"

    save_document(transport_path, lambda doc: create_transport_offer(doc, transport_supplier, commodities["RC1401"]))
    save_document(lab_path, lambda doc: create_lab_quotation(doc, lab_supplier, commodities["RC1202"]))
    save_document(
        digital_path,
        lambda doc: create_digital_cover_quote(doc, digital_supplier, commodities["RC1101"], commodities["RC1103"]),
    )

    metadata = [
        (transport_path, "OFF-A9001 synthetic transport offer", transport_supplier.supplier_name),
        (lab_path, "LAB-2401 synthetic laboratory quotation", lab_supplier.supplier_name),
        (digital_path, "8803 synthetic digital services quotation", digital_supplier.supplier_name),
    ]
    for path, title, author in metadata:
        doc = fitz.open(path)
        add_document_metadata(doc, title, author)
        doc.save(path.with_suffix(".tmp.pdf"), garbage=4, deflate=True)
        doc.close()
        path.with_suffix(".tmp.pdf").replace(path)
        print(path)


if __name__ == "__main__":
    main()
