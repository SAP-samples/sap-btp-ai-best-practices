"""
Excel extraction module for the Knowledge Graph Creation Pipeline.

This module handles Excel text extraction with enhanced metadata tracking,
converting tables to Markdown format and preserving sheet boundaries.
Part of Phase 1: Raw Text Extraction & Pre-processing.
"""

import re
import csv
import os
from typing import Dict, Any, List
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from dotenv import load_dotenv
from datetime import datetime

from ..models.kg_schema import SourceMetadata

# Load environment variables
load_dotenv()

# Configuration constants for table detection
DEFAULT_EMPTY_ROW_TOLERANCE = 3  # Number of consecutive empty rows that separate tables


def extract_excel_content(file_path: str, method: str = "openpyxl", fast_mode: bool = True, export_csv: bool = False, empty_row_tolerance: int = DEFAULT_EMPTY_ROW_TOLERANCE, include_metadata: bool = True) -> Dict[str, Any]:
    """
    Extract complete content from an Excel document with enhanced metadata tracking.
    
    This function implements Phase 1 of the KG Creation Pipeline, converting
    Excel files into structured Markdown text with preserved sheet boundaries.
    
    Args:
        file_path: Path to the Excel file
        method: Extraction method ("openpyxl" or "pandas")
        fast_mode: If True, skip complex region detection for faster processing
        export_csv: If True, export each table as a separate CSV file
        empty_row_tolerance: Number of consecutive empty rows to tolerate within a table
        include_metadata: Whether to include detailed metadata for each sheet
        
    Returns:
        Dictionary with extracted content and metadata:
        - filename: Name of the source file
        - file_path: Full path to the file
        - full_text: Complete extracted text with sheet headers
        - page_count: Number of sheets (treated as pages)
        - word_count: Total word count
        - extraction_method: Method used for extraction
        - pages: List of per-sheet content with metadata
        - extraction_timestamp: When the extraction occurred
        - source_metadata: List of SourceMetadata objects for each sheet
        
    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If the file is not an Excel file or cannot be processed
    """
    file_path = Path(file_path)
    
    # Validate file exists and is Excel
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
    
    if file_path.suffix.lower() not in ['.xlsx', '.xlsm', '.xls', '.xlsb']:
        raise ValueError(f"File must be Excel format, got: {file_path.suffix}")
    
    print(f"Extracting complete content from {file_path.name} using {method}...")
    
    result = {
        "filename": file_path.name,
        "file_path": str(file_path),
        "full_text": "",
        "page_count": 0,  # Will represent sheet count
        "word_count": 0,
        "extraction_method": method,
        "pages": [],  # Store per-sheet content (sheets = pages)
        "csv_exported": 0 if export_csv else None,  # Count of CSV files exported
        "extraction_timestamp": datetime.utcnow().isoformat(),
        "source_metadata": []  # List of SourceMetadata objects
    }
    
    try:
        if method == "openpyxl":
            result = _extract_with_openpyxl(file_path, result, fast_mode, export_csv, empty_row_tolerance, include_metadata)
        elif method == "pandas":
            result = _extract_with_pandas(file_path, result, export_csv, include_metadata)
        else:
            raise ValueError(f"Unknown extraction method: {method}")
            
    except Exception as e:
        raise ValueError(f"Error processing Excel file {file_path}: {str(e)}")
    
    # Validate extraction
    if not result["full_text"].strip():
        raise ValueError(f"No text content extracted from {file_path}")
    
    # Calculate statistics
    print("\nCalculating word count...")
    result["word_count"] = len(re.findall(r'\b\w+\b', result["full_text"]))
    
    print(f"\nExtraction complete: {result['page_count']} sheets, {result['word_count']} words")
    
    return result


def _extract_with_openpyxl(file_path: Path, result: Dict[str, Any], fast_mode: bool = True, export_csv: bool = False, empty_row_tolerance: int = DEFAULT_EMPTY_ROW_TOLERANCE, include_metadata: bool = True) -> Dict[str, Any]:
    """Extract text using openpyxl with enhanced formatting and metadata preservation."""
    
    print("Loading workbook...")
    # Load workbook with data_only=True to get calculated values instead of formulas
    wb = load_workbook(filename=str(file_path), data_only=True, read_only=True)
    result["page_count"] = len(wb.sheetnames)
    print(f"Found {result['page_count']} sheets")
    
    # Setup CSV export folder if needed
    csv_folder = None
    if export_csv:
        csv_folder = file_path.parent / "extracted_csv"
        print(f"CSV files will be saved to: {csv_folder}")
    
    all_text_parts = []
    
    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        print(f"\nProcessing sheet {sheet_idx + 1}/{result['page_count']}: {sheet_name}")
        sheet = wb[sheet_name]
        
        # Add sheet separator (similar to page separator in PDF) - using Markdown headers
        if sheet_idx > 0:
            all_text_parts.append(f"\n\n# Sheet {sheet_idx + 1}: {sheet_name}\n\n")
        else:
            all_text_parts.append(f"# Sheet 1: {sheet_name}\n\n")
        
        # Extract sheet content
        print(f"  Extracting content from sheet '{sheet_name}'...")
        sheet_content, table_count, csv_count = _extract_sheet_content(sheet, sheet_name, fast_mode, export_csv, csv_folder, empty_row_tolerance)
        print(f"  Found {table_count} table(s), {len(sheet_content)} characters")
        if export_csv and csv_count > 0:
            print(f"  Exported {csv_count} CSV file(s)")
            result["csv_exported"] += csv_count
        all_text_parts.append(sheet_content)
        
        # Store per-sheet content with enhanced metadata
        sheet_data = {
            "page_number": sheet_idx + 1,
            "sheet_name": sheet_name,
            "text": sheet_content,
            "character_count": len(sheet_content),
            "word_count": len(re.findall(r'\b\w+\b', sheet_content)),
            "table_count": table_count
        }
        
        if include_metadata:
            # Create SourceMetadata object for this sheet
            source_meta = SourceMetadata(
                filename=result["filename"],
                chunk_id=f"sheet_{sheet_name.replace(' ', '_')}"
            )
            result["source_metadata"].append(source_meta)
        
        result["pages"].append(sheet_data)
    
    wb.close()
    result["full_text"] = "".join(all_text_parts)
    return result


def _extract_with_pandas(file_path: Path, result: Dict[str, Any], export_csv: bool = False, include_metadata: bool = True) -> Dict[str, Any]:
    """Extract text using pandas with metadata tracking."""
    
    # Read all sheets
    excel_file = pd.ExcelFile(str(file_path))
    result["page_count"] = len(excel_file.sheet_names)
    
    # Setup CSV export folder if needed
    csv_folder = None
    if export_csv:
        csv_folder = file_path.parent / "extracted_csv"
        print(f"CSV files will be saved to: {csv_folder}")
    
    all_text_parts = []
    
    for sheet_idx, sheet_name in enumerate(excel_file.sheet_names):
        # Add sheet separator - using Markdown headers
        if sheet_idx > 0:
            all_text_parts.append(f"\n\n# Sheet {sheet_idx + 1}: {sheet_name}\n\n")
        else:
            all_text_parts.append(f"# Sheet 1: {sheet_name}\n\n")
        
        # Read sheet without headers to preserve all data
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
        
        # Convert DataFrame to table format
        if not df.empty:
            # Remove completely empty rows and columns
            df = df.dropna(how='all', axis=0).dropna(how='all', axis=1)
            
            if not df.empty:
                # Convert to list of lists for formatting
                table_data = df.fillna('').astype(str).values.tolist()
                table_title = f"Data from {sheet_name}"
                table_text = _format_table_as_markdown(table_data, table_title)
                all_text_parts.append(table_text)
                table_count = 1
                
                # Export to CSV if requested
                csv_count = 0
                if export_csv and csv_folder:
                    if _export_table_to_csv(table_data, table_title, csv_folder):
                        csv_count = 1
                        result["csv_exported"] += 1
            else:
                all_text_parts.append("(Empty sheet)")
                table_count = 0
                csv_count = 0
        else:
            all_text_parts.append("(Empty sheet)")
            table_count = 0
            csv_count = 0
        
        sheet_content = all_text_parts[-1]
        
        # Store per-sheet content with enhanced metadata
        sheet_data = {
            "page_number": sheet_idx + 1,
            "sheet_name": sheet_name,
            "text": sheet_content,
            "character_count": len(sheet_content),
            "word_count": len(re.findall(r'\b\w+\b', sheet_content)),
            "table_count": table_count
        }
        
        if include_metadata:
            # Create SourceMetadata object for this sheet
            source_meta = SourceMetadata(
                filename=result["filename"],
                chunk_id=f"sheet_{sheet_name.replace(' ', '_')}"
            )
            result["source_metadata"].append(source_meta)
        
        result["pages"].append(sheet_data)
    
    excel_file.close()
    result["full_text"] = "".join(all_text_parts)
    return result


def _extract_sheet_content(sheet, sheet_name: str, fast_mode: bool = True, export_csv: bool = False, csv_folder: Path = None, empty_row_tolerance: int = DEFAULT_EMPTY_ROW_TOLERANCE) -> tuple[str, int, int]:
    """Extract content from a single sheet, detecting tables and data regions."""
    
    if sheet.max_row == 0 or sheet.max_column == 0:
        return "(Empty sheet)", 0, 0
    
    # Find actual data boundaries
    min_row = sheet.min_row or 1
    max_row = sheet.max_row or 1
    min_col = sheet.min_column or 1
    max_col = sheet.max_column or 1
    
    print(f"    Sheet dimensions: {max_row} rows x {max_col} columns")
    
    if fast_mode:
        print(f"    Fast mode: treating entire sheet as one table")
        # In fast mode, just treat the entire sheet as one table
        data_regions = [(min_row, max_row, min_col, max_col)]
        merged_ranges = []
    else:
        # Detect merged cell ranges
        merged_ranges = list(sheet.merged_cells.ranges) if hasattr(sheet, 'merged_cells') else []
        print(f"    Found {len(merged_ranges)} merged cell ranges")
        
        # Detect data regions (simple approach: continuous non-empty areas)
        print(f"    Detecting data regions...")
        data_regions = _detect_data_regions(sheet, min_row, max_row, min_col, max_col, empty_row_tolerance)
        print(f"    Found {len(data_regions)} data region(s)")
        
        if not data_regions:
            # If no regions detected, treat entire sheet as one table
            data_regions = [(min_row, max_row, min_col, max_col)]
    
    content_parts = []
    table_count = 0
    csv_exported_count = 0
    
    for region_idx, (r_min_row, r_max_row, r_min_col, r_max_col) in enumerate(data_regions):
        region_rows = r_max_row - r_min_row + 1
        region_cols = r_max_col - r_min_col + 1
        print(f"    Processing region {region_idx + 1}: {region_rows}x{region_cols} cells")
        table_data = []
        
        for row in range(r_min_row, r_max_row + 1):
            row_data = []
            for col in range(r_min_col, r_max_col + 1):
                cell = sheet.cell(row=row, column=col)
                
                # Check if cell is part of a merged range
                cell_value = ""
                is_merged = False
                
                for merged_range in merged_ranges:
                    min_mr, min_mc, max_mr, max_mc = range_boundaries(str(merged_range))
                    if min_mr <= row <= max_mr and min_mc <= col <= max_mc:
                        is_merged = True
                        # Only show value for top-left cell of merged range
                        if row == min_mr and col == min_mc:
                            cell_value = str(cell.value) if cell.value is not None else ""
                            # Add merge indicator
                            if max_mr > min_mr or max_mc > min_mc:
                                cell_value = f"[MERGED: {cell_value}]"
                        break
                
                if not is_merged:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    
                    # Preserve some formatting
                    if cell.font and hasattr(cell.font, 'bold') and cell.font.bold:
                        cell_value = f"**{cell_value}**" if cell_value else ""
                
                row_data.append(cell_value)
            
            table_data.append(row_data)
        
        # Format table
        if table_data and any(any(cell for cell in row) for row in table_data):
            table_count += 1
            if len(data_regions) > 1:
                table_title = f"Table {table_count} from {sheet_name}"
            else:
                table_title = f"Data from {sheet_name}"
            
            formatted_table = _format_table_as_markdown(table_data, table_title)
            content_parts.append(formatted_table)
            
            # Export to CSV if requested
            if export_csv and csv_folder:
                if _export_table_to_csv(table_data, table_title, csv_folder):
                    csv_exported_count += 1
    
    return "\n\n".join(content_parts) if content_parts else "(No data found)", table_count, csv_exported_count


def _detect_data_regions(sheet, min_row: int, max_row: int, min_col: int, max_col: int, empty_row_tolerance: int = DEFAULT_EMPTY_ROW_TOLERANCE) -> list:
    """Detect separate data regions/tables in a sheet.
    
    Args:
        sheet: The worksheet object
        min_row: Minimum row number to scan
        max_row: Maximum row number to scan
        min_col: Minimum column number to scan
        max_col: Maximum column number to scan
        empty_row_tolerance: Number of consecutive empty rows to tolerate within a table
    
    Returns:
        List of tuples (min_row, max_row, min_col, max_col) representing data regions
    """
    
    # Simple detection: find continuous non-empty regions
    # This is a basic implementation - could be enhanced with more sophisticated detection
    
    regions = []
    visited = set()
    
    total_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
    print(f"      Scanning {total_cells} cells for data regions...")
    
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if (row, col) in visited:
                continue
                
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                # Found a non-empty cell, explore the region
                region_bounds = _explore_region(sheet, row, col, max_row, max_col, visited, empty_row_tolerance)
                if region_bounds:
                    regions.append(region_bounds)
    
    return regions


def _explore_region(sheet, start_row: int, start_col: int, max_row: int, max_col: int, visited: set, empty_row_tolerance: int = DEFAULT_EMPTY_ROW_TOLERANCE) -> tuple:
    """Explore a continuous data region starting from a non-empty cell.
    
    Args:
        sheet: The worksheet object
        start_row: Starting row for exploration
        start_col: Starting column for exploration
        max_row: Maximum row boundary
        max_col: Maximum column boundary
        visited: Set of already visited cells
        empty_row_tolerance: Number of consecutive empty rows before considering it a table boundary
    
    Returns:
        Tuple (min_row, max_row, min_col, max_col) representing the region bounds
    """
    
    # Find bounds of continuous data region
    r_min = start_row
    r_max = start_row
    c_min = start_col
    c_max = start_col
    
    # Simple approach: find rectangle bounds with tolerance for empty cells
    # empty_row_tolerance is now a parameter with default value 2
    
    # Expand right and down to find bounds
    last_non_empty_row = start_row
    for row in range(start_row, min(start_row + 100, max_row + 1)):  # Limit to 100 rows per region
        row_has_data = False
        for col in range(start_col, min(start_col + 50, max_col + 1)):  # Limit to 50 columns per region
            cell = sheet.cell(row=row, column=col)
            visited.add((row, col))
            
            if cell.value is not None:
                row_has_data = True
                r_max = max(r_max, row)
                c_max = max(c_max, col)
                c_min = min(c_min, col)
                last_non_empty_row = row
        
        if not row_has_data and row - last_non_empty_row > empty_row_tolerance:
            break
    
    # Expand left and up to find full bounds
    for row in range(r_min, r_max + 1):
        for col in range(max(1, c_min - 10), c_max + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                c_min = min(c_min, col)
                visited.add((row, col))
    
    # Mark all cells in region as visited
    for row in range(r_min, r_max + 1):
        for col in range(c_min, c_max + 1):
            visited.add((row, col))
    
    return (r_min, r_max, c_min, c_max)


def _format_table_as_markdown(table: list, table_title: str = "Table") -> str:
    """
    Format a table as Markdown table.
    Converts Excel table data to proper Markdown table format.
    """
    
    if not table:
        return ""
    
    # Remove None values and convert to strings, escape markdown characters
    clean_table = []
    for row in table:
        clean_row = []
        for cell in row:
            cell_str = str(cell) if cell is not None else ""
            # Escape pipe characters in cell content
            cell_str = cell_str.replace("|", "\\|")
            clean_row.append(cell_str)
        clean_table.append(clean_row)
    
    if not clean_table:
        return ""
    
    # Ensure all rows have the same number of columns
    max_cols = max(len(row) for row in clean_table)
    for row in clean_table:
        while len(row) < max_cols:
            row.append("")
    
    # Format as markdown table with proper spacing
    formatted_lines = [f"## {table_title}\n"]
    
    # Calculate column widths for better formatting
    col_widths = []
    for col_idx in range(max_cols):
        max_width = 3  # Minimum width of 3 for '---'
        for row in clean_table:
            if col_idx < len(row):
                max_width = max(max_width, len(row[col_idx]))
        col_widths.append(max_width)
    
    for row_num, row in enumerate(clean_table):
        # Format row with markdown table syntax and proper alignment
        cells = []
        for col_idx in range(max_cols):
            cell = row[col_idx] if col_idx < len(row) else ""
            # Pad cell content to column width for better alignment
            cells.append(cell.ljust(col_widths[col_idx]))
        
        markdown_row = "| " + " | ".join(cells) + " |"
        formatted_lines.append(markdown_row)
        
        # Add separator after header (first row) with proper column widths
        if row_num == 0 and len(clean_table) > 1:
            separator_cells = ["-" * width for width in col_widths]
            separator = "| " + " | ".join(separator_cells) + " |"
            formatted_lines.append(separator)
    
    return "\n".join(formatted_lines)


def _format_table_as_text(table: list, table_title: str = "Table") -> str:
    """
    Format a table as readable text.
    This is the same function as in pdf_processor.py to ensure consistent output.
    Kept for backward compatibility.
    """
    
    if not table:
        return ""
    
    # Remove None values and convert to strings
    clean_table = []
    for row in table:
        clean_row = [str(cell) if cell is not None else "" for cell in row]
        clean_table.append(clean_row)
    
    if not clean_table:
        return ""
    
    # Calculate column widths
    max_cols = max(len(row) for row in clean_table)
    col_widths = [0] * max_cols
    
    for row in clean_table:
        for i, cell in enumerate(row):
            if i < max_cols:
                col_widths[i] = max(col_widths[i], len(cell))
    
    # Format table
    formatted_lines = [f"=== {table_title} ==="]
    
    for row_num, row in enumerate(clean_table):
        formatted_row = []
        for i in range(max_cols):
            cell = row[i] if i < len(row) else ""
            formatted_row.append(cell.ljust(col_widths[i]))
        
        formatted_lines.append(" | ".join(formatted_row))
        
        # Add separator after header
        if row_num == 0 and len(clean_table) > 1:
            separator = " | ".join(["-" * width for width in col_widths])
            formatted_lines.append(separator)
    
    return "\n".join(formatted_lines)


def sanitize_filename(name: str) -> str:
    """
    Sanitize a string to be safe for use as a filename.
    Replaces invalid characters and ensures reasonable length.
    """
    # Replace spaces with underscores
    name = name.replace(" ", "_")
    
    # Remove or replace invalid filename characters
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, "_")
    
    # Remove multiple consecutive underscores
    while "__" in name:
        name = name.replace("__", "_")
    
    # Remove leading/trailing underscores
    name = name.strip("_")
    
    # Ensure reasonable length (max 200 characters for filename)
    if len(name) > 200:
        name = name[:200]
    
    # Ensure it's not empty
    if not name:
        name = "table"
    
    return name


def _export_table_to_csv(table_data: list, table_title: str, csv_folder: Path) -> bool:
    """
    Export table data to a CSV file.
    
    Args:
        table_data: List of lists containing table data
        table_title: Title/name for the table (used for filename)
        csv_folder: Path to the CSV output folder
        
    Returns:
        bool: True if export successful, False otherwise
    """
    if not table_data:
        return False
    
    try:
        # Create CSV folder if it doesn't exist
        csv_folder.mkdir(exist_ok=True)
        
        # Sanitize filename
        safe_filename = sanitize_filename(table_title)
        csv_file = csv_folder / f"{safe_filename}.csv"
        
        # Write CSV file
        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Write all rows
            for row in table_data:
                # Clean up the row data
                clean_row = []
                for cell in row:
                    cell_str = str(cell) if cell is not None else ""
                    # Remove markdown formatting for CSV
                    cell_str = cell_str.replace("**", "")  # Remove bold markers
                    cell_str = cell_str.replace("[MERGED: ", "").replace("]", "")  # Clean merged cell markers
                    clean_row.append(cell_str)
                
                writer.writerow(clean_row)
        
        print(f"    CSV exported: {csv_file.name}")
        return True
        
    except Exception as e:
        print(f"    Error exporting CSV for '{table_title}': {e}")
        return False


# Additional utility functions matching pdf_processor.py interface



def extract_excel_for_kg_pipeline(file_path: str) -> Dict[str, Any]:
    """
    Specialized extraction function for the KG Creation Pipeline.
    
    This is a convenience wrapper that ensures all necessary metadata
    is extracted for downstream processing phases.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Extraction result optimized for KG pipeline processing
    """
    return extract_excel_content(
        file_path=file_path,
        method="openpyxl",  # Default to openpyxl for better control
        fast_mode=True,  # Fast mode for efficiency
        export_csv=False,  # No CSV export by default
        include_metadata=True
    )


def get_sheet_separators(result: Dict[str, Any]) -> List[str]:
    """
    Extract the sheet header patterns from the full text.
    
    This is useful for the chunking phase to split the document
    by sheet boundaries.
    
    Args:
        result: Extraction result dictionary
        
    Returns:
        List of sheet header strings found in the text
    """
    import re
    
    # Find all sheet headers in the text
    header_pattern = r'# Sheet \d+: [^\n]+'
    headers = re.findall(header_pattern, result['full_text'])
    
    return headers


if __name__ == "__main__":
    # Example usage
    import sys
    FAST_MODE = True  # Set to True for faster extraction with less detail
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = "data/sample.xlsx"
    
    try:
        # Extract content using the KG pipeline function
        result = extract_excel_for_kg_pipeline(excel_file)
        
        print("\n" + "="*50)
        print(f"Extraction complete for: {result['filename']}")
        print(f"Sheets: {result['page_count']}")
        print(f"Total words: {result['word_count']}")
        print(f"Source metadata objects: {len(result['source_metadata'])}")
        
        # Save to markdown file
        output_file = Path(excel_file).stem + "_extracted.md"
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(result['full_text'])
        print(f"\nMarkdown saved to: {output_file}")
        
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)